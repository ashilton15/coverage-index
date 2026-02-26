"""Coverage Index - Media Coverage Scoring Tool by Agnostic."""

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import time
import json

from config import SCORING_MODEL, COLORS, get_score_color, get_score_grade
from scraper import scrape_url, extract_domain
from scorer import create_client, score_article, lookup_outlet
from sheets import (
    load_outlets_from_excel,
    connect_to_sheets,
    get_or_create_spreadsheet,
    load_clients_from_sheet,
    save_client_to_sheet,
    delete_client_from_sheet,
    load_outlets_from_sheet,
    save_outlets_to_sheet,
)
from pdf_report import generate_coverage_report
import os


@st.cache_resource
def get_sheets_spreadsheet():
    """Connect to Google Sheets and return the CoverageIndex spreadsheet.

    Cached so the connection is reused across reruns.
    Returns None if credentials are not configured.
    """
    try:
        creds = dict(st.secrets["gcp_service_account"])
        gc = connect_to_sheets(creds)
        spreadsheet = get_or_create_spreadsheet(gc, "Coverage Scorer Data")
        return spreadsheet
    except Exception:
        return None

# Page config
st.set_page_config(
    page_title="Coverage Index",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# Custom CSS for black/white theme
st.markdown("""
<style>
    .stApp { background-color: #FAFAFA; }
    h1, h2, h3 { color: #000000 !important; }

    .big-score {
        font-size: 64px;
        font-weight: bold;
        color: #000000;
        line-height: 1;
    }
    .grade-badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 4px;
        font-weight: bold;
        font-size: 18px;
    }
    .tier-bar {
        background: #E5E5E5;
        border-radius: 4px;
        height: 8px;
        margin: 4px 0;
    }
    .tier-fill {
        height: 100%;
        border-radius: 4px;
    }
    .penalty-tag {
        color: #EF4444;
        font-size: 12px;
    }
    .failed-article {
        border-left: 4px solid #EF4444;
        padding-left: 12px;
        margin: 8px 0;
    }
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


def init_session_state():
    """Initialize session state variables."""
    defaults = {
        "authenticated": False,
        "current_tab": "analyze",
        "step": 1,
        "selected_client": None,
        "selected_campaign": None,
        "articles": [],
        "results": [],
        "failed_articles": [],
        "outlets_df": None,
        "clients": {},  # Dict of client_name -> client data with campaigns
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def check_auth():
    """Check if user is authenticated."""
    if st.session_state.authenticated:
        return True

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        try:
            st.image("CI - logo.png", width=200)
        except Exception:
            st.markdown("# 📊 Coverage Index")

        st.markdown("### Enter password to continue")
        password = st.text_input("Password:", type="password", label_visibility="collapsed")

        if password:
            if password == st.secrets.get("app_password", "coverageindex"):
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Incorrect password")

    return False


def load_data():
    """Load outlets and clients data from Google Sheets (with local fallback for outlets)."""
    spreadsheet = get_sheets_spreadsheet()

    # Load outlets
    if st.session_state.outlets_df is None:
        loaded = False
        # Try Google Sheets first
        if spreadsheet is not None:
            try:
                df = load_outlets_from_sheet(spreadsheet)
                if not df.empty:
                    st.session_state.outlets_df = df
                    loaded = True
            except Exception:
                pass
        # Fall back to local Excel file
        if not loaded:
            try:
                st.session_state.outlets_df = load_outlets_from_excel("Media Outlets.xlsx")
                # Sync local outlets to Sheets on first load
                if spreadsheet is not None and st.session_state.outlets_df is not None and not st.session_state.outlets_df.empty:
                    try:
                        save_outlets_to_sheet(spreadsheet, st.session_state.outlets_df)
                    except Exception:
                        pass
            except Exception as e:
                st.error(f"Could not load outlets: {e}")
                st.session_state.outlets_df = pd.DataFrame()

    # Load clients from Google Sheets
    if not st.session_state.clients:
        if spreadsheet is not None:
            try:
                clients = load_clients_from_sheet(spreadsheet)
                if clients:
                    st.session_state.clients = clients
            except Exception:
                pass
        # Fall back to default if nothing loaded
        if not st.session_state.clients:
            st.session_state.clients = {"Coinbase": get_default_coinbase_client()}
            # Save default client to Sheets
            if spreadsheet is not None:
                try:
                    save_client_to_sheet(spreadsheet, "Coinbase", st.session_state.clients["Coinbase"])
                except Exception:
                    pass


def get_default_coinbase_client():
    """Return default Coinbase client configuration with campaigns."""
    return {
        "name": "Coinbase",
        "industry": "Cryptocurrency / Fintech",
        "campaigns": {
            "Stablecoin Regulation": {
                "name": "Stablecoin Regulation",
                "spokespeople": [
                    "Lucas Matheson",
                    "Tom Duff Gordon",
                ],
                "key_messages_high": [
                    "Canada's new stablecoin framework is a positive first step towards positioning the country as a digital economy leader",
                    "A competitive Canadian stablecoin regime requires Bank of Canada-anchored oversight and fast-tracked interim approval for CAD stablecoins",
                    "The framework signals Canada is ready to lead again on financial innovation",
                ],
                "key_messages_medium": [
                    "Canada is the only G7 country without a well-defined regulatory system for stablecoins",
                    "Stablecoin legislation is not a partisan issue - it's a national priority",
                    "Building a Canadian stablecoin framework protects our financial sovereignty",
                ],
                "key_messages_low": [
                    "Stablecoins can make Canada's financial system safer, fairer, and more inclusive",
                    "A well-regulated stablecoin system can make life more affordable for Canadians",
                    "Without clear rules, Canada risks becoming dependent on U.S.-issued stablecoins",
                    "Coinbase has worked closely with Canadian regulators for years",
                ],
                "competitors": ["Binance", "Kraken", "Crypto.com"],
            }
        }
    }


def render_navigation():
    """Render the main navigation tabs."""
    col1, col2, col3 = st.columns([1, 3, 1])

    with col1:
        try:
            st.image("CI - logo.png", width=80)
        except Exception:
            st.markdown("### 📊")

    with col2:
        tab1, tab2, tab3 = st.tabs(["📊 Analyze Coverage", "👥 Client Profiles", "🔧 Tools"])

        if tab1:
            st.session_state.current_tab = "analyze"
        if tab2:
            st.session_state.current_tab = "profiles"
        if tab3:
            st.session_state.current_tab = "tools"

    return st.session_state.current_tab


def render_tools():
    """Render the tools tab with utilities."""
    st.markdown("## Tools")

    st.markdown("### 🗑️ Clear Score Cache")
    st.markdown("Scores are cached for consistency. Clear the cache to re-analyze articles from scratch.")

    if st.button("Clear Cache", type="primary"):
        from scorer import clear_score_cache
        clear_score_cache()
        st.success("✅ Score cache cleared. Articles will be re-scored on next analysis.")

    st.markdown("---")

    st.markdown("### 📊 Scoring Model Info")
    st.markdown("""
    **Current scoring weights:**
    - **Tier 1 - Foundational:** 39 pts (outlet tier, sentiment, article type, tone)
    - **Tier 2 - Messaging:** 35 pts (key messages, quotes, framing)
    - **Tier 3 - Prominence:** 5 pts (position in article)
    - **Tier 4 - Credibility:** 10 pts (spokesperson, expertise framing)
    - **Tier 5 - Competitive:** 9 pts (vs competitors)
    - **Tier 6 - Audience Fit:** 6.5 pts (outlet relevance)
    - **Tier 7 - Supporting:** 2.5 pts (CTA, length, exclusivity)
    - **Bonus:** up to 20 pts (op-ed +12, headline +3, etc.)

    **Score floors (minimum guaranteed scores):**
    - Tier 1 outlet + client op-ed = minimum **85**
    - Tier 1 + op-ed + all key messages = minimum **90**
    - Tier 1 + spokesperson quoted + 2+ key messages = minimum **70**
    """)


def render_client_profiles():
    """Render the client profiles management tab."""
    st.markdown("## Client Profiles")
    st.markdown("Manage clients and their campaign configurations.")

    # Client selector
    client_names = list(st.session_state.clients.keys())

    col1, col2 = st.columns([3, 1])
    with col1:
        selected_client = st.selectbox("Select client:", options=[""] + client_names, key="profile_client_select")
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("➕ New Client"):
            st.session_state.show_new_client_form = True

    # New client form
    if st.session_state.get("show_new_client_form"):
        render_new_client_form()
        return

    if not selected_client:
        st.info("Select a client to view or edit their campaigns.")
        return

    client = st.session_state.clients[selected_client]

    col_title, col_delete = st.columns([4, 1])
    with col_title:
        st.markdown(f"### {client['name']}")
        st.markdown(f"**Industry:** {client.get('industry', 'Not set')}")
    with col_delete:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑️ Delete Client", key="delete_client_btn"):
            # Remove from session state
            del st.session_state.clients[selected_client]
            # Remove from Google Sheets
            spreadsheet = get_sheets_spreadsheet()
            if spreadsheet is not None:
                try:
                    delete_client_from_sheet(spreadsheet, selected_client)
                except Exception:
                    pass
            st.rerun()

    # Campaign management
    st.markdown("---")
    st.markdown("### Campaigns")

    campaigns = client.get("campaigns", {})

    col1, col2 = st.columns([3, 1])
    with col1:
        campaign_names = list(campaigns.keys())
        selected_campaign = st.selectbox("Select campaign:", options=[""] + campaign_names, key="profile_campaign_select")
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("➕ New Campaign"):
            st.session_state.editing_campaign = {"client": selected_client, "campaign": None}

    # Edit campaign form
    if st.session_state.get("editing_campaign"):
        render_campaign_form(st.session_state.editing_campaign["client"], st.session_state.editing_campaign.get("campaign"))
        return

    if selected_campaign and selected_campaign in campaigns:
        campaign = campaigns[selected_campaign]
        render_campaign_details(selected_client, campaign)


def render_new_client_form():
    """Render form for creating a new client."""
    st.markdown("### New Client")

    with st.form("new_client_form"):
        name = st.text_input("Client name:")
        industry = st.text_input("Industry:")

        col1, col2 = st.columns(2)
        with col1:
            if st.form_submit_button("Create Client", type="primary"):
                if name:
                    client_data = {
                        "name": name,
                        "industry": industry,
                        "campaigns": {}
                    }
                    st.session_state.clients[name] = client_data
                    # Persist to Google Sheets
                    spreadsheet = get_sheets_spreadsheet()
                    if spreadsheet is not None:
                        try:
                            save_client_to_sheet(spreadsheet, name, client_data)
                        except Exception:
                            pass
                    st.session_state.show_new_client_form = False
                    st.rerun()
        with col2:
            if st.form_submit_button("Cancel"):
                st.session_state.show_new_client_form = False
                st.rerun()


def render_campaign_form(client_name: str, existing_campaign: dict = None):
    """Render form for creating or editing a campaign."""
    st.markdown("### " + ("Edit Campaign" if existing_campaign else "New Campaign"))

    with st.form("campaign_form"):
        name = st.text_input("Campaign name:", value=existing_campaign.get("name", "") if existing_campaign else "")

        # Support for multiple spokespeople
        existing_spokespeople = existing_campaign.get("spokespeople", []) if existing_campaign else []
        # Backwards compatibility: check old "spokesperson" field
        if not existing_spokespeople and existing_campaign and existing_campaign.get("spokesperson"):
            existing_spokespeople = [existing_campaign.get("spokesperson")]

        spokespeople = st.text_area(
            "Spokespeople (one per line):",
            value="\n".join(existing_spokespeople),
            height=80,
            help="Enter each spokesperson's full name on a separate line. Any of these can be quoted or write op-eds."
        )

        st.markdown("#### Key Messages")
        st.caption("Organize messages by priority. HIGH = 3 pts, MEDIUM = 1.5 pts, LOW = 0.5 pts")

        high = st.text_area(
            "HIGH priority (one per line):",
            value="\n".join(existing_campaign.get("key_messages_high", [])) if existing_campaign else "",
            height=100,
        )
        medium = st.text_area(
            "MEDIUM priority (one per line):",
            value="\n".join(existing_campaign.get("key_messages_medium", [])) if existing_campaign else "",
            height=100,
        )
        low = st.text_area(
            "LOW priority (one per line):",
            value="\n".join(existing_campaign.get("key_messages_low", [])) if existing_campaign else "",
            height=100,
        )

        competitors = st.text_input(
            "Competitors (comma separated):",
            value=", ".join(existing_campaign.get("competitors", [])) if existing_campaign else "",
        )

        col1, col2 = st.columns(2)
        with col1:
            if st.form_submit_button("Save Campaign", type="primary"):
                campaign = {
                    "name": name,
                    "spokespeople": [s.strip() for s in spokespeople.split("\n") if s.strip()],
                    "key_messages_high": [m.strip() for m in high.split("\n") if m.strip()],
                    "key_messages_medium": [m.strip() for m in medium.split("\n") if m.strip()],
                    "key_messages_low": [m.strip() for m in low.split("\n") if m.strip()],
                    "competitors": [c.strip() for c in competitors.split(",") if c.strip()],
                }

                st.session_state.clients[client_name]["campaigns"][name] = campaign
                # Persist to Google Sheets
                spreadsheet = get_sheets_spreadsheet()
                if spreadsheet is not None:
                    try:
                        save_client_to_sheet(spreadsheet, client_name, st.session_state.clients[client_name])
                    except Exception:
                        pass
                st.session_state.editing_campaign = None
                st.rerun()

        with col2:
            if st.form_submit_button("Cancel"):
                st.session_state.editing_campaign = None
                st.rerun()


def get_spokespeople(campaign: dict) -> list:
    """Get list of spokespeople from campaign (handles old and new format)."""
    spokespeople = campaign.get("spokespeople", [])
    # Backwards compatibility
    if not spokespeople and campaign.get("spokesperson"):
        spokespeople = [campaign.get("spokesperson")]
    return spokespeople


def render_campaign_details(client_name: str, campaign: dict):
    """Render details of a campaign."""
    st.markdown(f"#### {campaign['name']}")
    spokespeople = get_spokespeople(campaign)
    st.markdown(f"**Spokespeople:** {', '.join(spokespeople) if spokespeople else 'Not set'}")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("**HIGH Priority Messages:**")
        for msg in campaign.get("key_messages_high", []):
            st.markdown(f"- {msg}")

    with col2:
        st.markdown("**MEDIUM Priority Messages:**")
        for msg in campaign.get("key_messages_medium", []):
            st.markdown(f"- {msg}")

    with col3:
        st.markdown("**LOW Priority Messages:**")
        for msg in campaign.get("key_messages_low", []):
            st.markdown(f"- {msg}")

    st.markdown(f"**Competitors:** {', '.join(campaign.get('competitors', [])) or 'None'}")

    if st.button("✏️ Edit Campaign"):
        st.session_state.editing_campaign = {"client": client_name, "campaign": campaign}
        st.rerun()


def render_analyze_tab():
    """Render the main analysis tab."""
    # Step indicator
    steps = ["Upload", "Select Campaign", "Analyze", "Results"]
    cols = st.columns(len(steps))
    for i, (col, step_name) in enumerate(zip(cols, steps)):
        step_num = i + 1
        with col:
            if step_num < st.session_state.step:
                st.markdown(f"✅ **{step_name}**")
            elif step_num == st.session_state.step:
                st.markdown(f"🔵 **{step_name}**")
            else:
                st.markdown(f"⚪ {step_name}")

    st.markdown("---")

    # Render current step
    if st.session_state.step == 1:
        step1_upload()
    elif st.session_state.step == 2:
        step2_select_campaign()
    elif st.session_state.step == 3:
        step3_analyze()
    elif st.session_state.step == 4:
        step4_results()


def step1_upload():
    """Step 1: Upload coverage for analysis."""
    st.markdown("## Step 1: Add Coverage")

    tab1, tab2 = st.tabs(["📁 Batch Upload", "🔗 Single Article"])

    with tab1:
        uploaded_file = st.file_uploader(
            "Upload Excel file with coverage URLs",
            type=["xlsx", "xls"],
            help="File should have headlines with hyperlinked URLs",
        )

        if uploaded_file:
            articles = parse_excel_upload(uploaded_file)
            if articles:
                st.success(f"Found {len(articles)} articles with URLs")

                with st.expander("Preview articles", expanded=True):
                    for i, article in enumerate(articles[:5]):
                        st.markdown(f"**{i+1}.** {article.get('headline', 'No headline')}")
                        st.caption(f"{article.get('publication', 'Unknown')} - {article.get('url', '')[:60]}...")
                    if len(articles) > 5:
                        st.caption(f"...and {len(articles) - 5} more")

                if st.button("Continue with these articles", type="primary"):
                    st.session_state.articles = articles
                    st.session_state.step = 2
                    st.rerun()

    with tab2:
        st.markdown("### Analyze a Single Article")

        # Initialize session state for single article fetch
        if "single_article_fetched" not in st.session_state:
            st.session_state.single_article_fetched = False
            st.session_state.single_article_data = {}

        # Step 1: Enter URL and fetch
        url = st.text_input("Enter article URL:", key="single_article_url")

        col1, col2 = st.columns([1, 3])
        with col1:
            fetch_clicked = st.button("🔍 Fetch Article", type="primary", disabled=not url)

        if fetch_clicked and url:
            with st.spinner("Fetching article..."):
                from scraper import scrape_url
                from scorer import lookup_outlet, detect_op_ed_url

                result = scrape_url(url)
                domain = extract_domain(url)

                # Look up outlet
                outlets_df = st.session_state.get("outlets_df")
                outlet_info = {"name": "", "tier": 3, "type": "Online"}
                if outlets_df is not None:
                    outlet_info = lookup_outlet(domain, outlets_df)

                st.session_state.single_article_fetched = True
                st.session_state.single_article_data = {
                    "url": url,
                    "fetch_success": result["success"],
                    "headline": result.get("title", ""),
                    "content": result.get("content", ""),
                    "word_count": result.get("word_count", 0),
                    "error": result.get("error", ""),
                    "domain": domain,
                    "outlet_name": outlet_info.get("name", ""),
                    "outlet_tier": outlet_info.get("tier", 3),
                    "outlet_found": outlet_info.get("found", False),
                    "is_op_ed": detect_op_ed_url(url),
                    "raw_html": result.get("raw_html"),
                    "author": result.get("author"),
                }
                st.rerun()

        # Step 2: Show fetched data and allow editing
        if st.session_state.single_article_fetched:
            data = st.session_state.single_article_data

            if data.get("fetch_success"):
                st.success(f"✅ Article fetched successfully! ({data.get('word_count', 0)} words)")
                if data.get("is_op_ed"):
                    st.info("📝 URL pattern suggests this is an **opinion/op-ed** piece.")
            else:
                st.warning(f"⚠️ Could not fetch article: {data.get('error', 'Unknown error')}")
                st.markdown("You can still analyze by filling in the fields manually below.")

            st.markdown("---")
            st.markdown("**Review and edit article details:**")

            # Editable fields
            headline = st.text_input("Headline:", value=data.get("headline", ""), key="single_headline")
            byline = st.text_input("Byline/Author:", placeholder="e.g., Tom Duff Gordon", key="single_byline")

            col1, col2 = st.columns(2)
            with col1:
                outlet_name = st.text_input("Outlet Name:", value=data.get("outlet_name", ""), key="single_outlet")
            with col2:
                tier_options = ["Auto-detect", "Tier 1", "Tier 2", "Tier 3", "Tier 4"]
                default_tier = f"Tier {data.get('outlet_tier', 3)}" if data.get("outlet_found") else "Auto-detect"
                outlet_tier = st.selectbox("Outlet Tier:", options=tier_options,
                                          index=tier_options.index(default_tier) if default_tier in tier_options else 0,
                                          key="single_tier")

            # Content preview/edit
            content = st.text_area(
                "Article Content:",
                value=data.get("content", ""),
                height=200,
                help="Review the extracted content. Edit if needed or paste full article text.",
                key="single_content"
            )

            st.markdown("---")

            # Action buttons
            col1, col2, col3 = st.columns([1, 1, 2])
            with col1:
                if st.button("🔄 Reset"):
                    st.session_state.single_article_fetched = False
                    st.session_state.single_article_data = {}
                    st.rerun()

            with col2:
                analyze_disabled = not (headline and content)
                if st.button("▶️ Run Analysis", type="primary", disabled=analyze_disabled):
                    # Store article data for analysis
                    st.session_state.articles = [{
                        "url": data.get("url", ""),
                        "headline": headline,
                        "publication": outlet_name,
                        "byline": byline,
                        "domain": data.get("domain", ""),
                        "content_override": content,  # Will use this instead of fetching again
                        "tier_override": int(outlet_tier.split()[1]) if outlet_tier != "Auto-detect" else None,
                        "raw_html": data.get("raw_html"),
                        "author": data.get("author"),
                    }]
                    st.session_state.step = 2
                    st.rerun()

            if analyze_disabled:
                st.caption("⚠️ Headline and content are required to run analysis.")


def parse_excel_upload(uploaded_file) -> list:
    """Parse uploaded Excel file to extract articles with hyperlinks."""
    try:
        wb = load_workbook(uploaded_file, data_only=False)
        ws = wb.active
        articles = []
        headers = [cell.value for cell in ws[1]]

        # Find columns
        headline_col = None
        pub_col = None
        date_col = None

        for i, header in enumerate(headers):
            if header:
                header_lower = str(header).lower()
                if "link" in header_lower or "url" in header_lower or "headline" in header_lower:
                    headline_col = i
                if "publication" in header_lower or "outlet" in header_lower:
                    pub_col = i
                if "date" in header_lower:
                    date_col = i

        if headline_col is None:
            headline_col = 2  # Default to column C

        for row_num in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=headline_col + 1)

            if not cell.value:
                continue

            headline = str(cell.value)
            url = None

            if cell.hyperlink:
                url = cell.hyperlink.target

            if not url:
                continue

            publication = ""
            if pub_col is not None:
                pub_cell = ws.cell(row=row_num, column=pub_col + 1)
                if pub_cell.value:
                    publication = str(pub_cell.value)

            date = ""
            if date_col is not None:
                date_cell = ws.cell(row=row_num, column=date_col + 1)
                if date_cell.value:
                    date = str(date_cell.value)

            articles.append({
                "url": url,
                "headline": headline,
                "publication": publication,
                "date": date,
                "domain": extract_domain(url),
            })

        return articles

    except Exception as e:
        st.error(f"Error parsing file: {e}")
        return []


def step2_select_campaign():
    """Step 2: Select client and campaign."""
    st.markdown("## Step 2: Select Client & Campaign")

    # Client selection
    client_names = list(st.session_state.clients.keys())

    if not client_names:
        st.warning("No clients configured. Go to Client Profiles to add one.")
        return

    selected_client_name = st.selectbox("Select client:", options=client_names)

    if selected_client_name:
        client = st.session_state.clients[selected_client_name]
        campaigns = client.get("campaigns", {})

        if not campaigns:
            st.warning(f"No campaigns configured for {selected_client_name}. Go to Client Profiles to add one.")
            return

        campaign_names = list(campaigns.keys())
        selected_campaign_name = st.selectbox("Select campaign:", options=campaign_names)

        if selected_campaign_name:
            campaign = campaigns[selected_campaign_name]

            with st.expander("Campaign details", expanded=True):
                spokespeople = get_spokespeople(campaign)
                st.markdown(f"**Spokespeople:** {', '.join(spokespeople) if spokespeople else 'Not set'}")

                col1, col2, col3 = st.columns(3)
                with col1:
                    st.markdown(f"**HIGH priority:** {len(campaign.get('key_messages_high', []))} messages")
                with col2:
                    st.markdown(f"**MEDIUM priority:** {len(campaign.get('key_messages_medium', []))} messages")
                with col3:
                    st.markdown(f"**LOW priority:** {len(campaign.get('key_messages_low', []))} messages")

                st.markdown(f"**Competitors:** {', '.join(campaign.get('competitors', [])) or 'None'}")

            col1, col2 = st.columns(2)
            with col1:
                if st.button("← Back"):
                    st.session_state.step = 1
                    st.rerun()
            with col2:
                if st.button("Continue to Analysis →", type="primary"):
                    st.session_state.selected_client = client
                    st.session_state.selected_campaign = campaign
                    st.session_state.step = 3
                    st.rerun()


def step3_analyze():
    """Step 3: Scrape and analyze articles."""
    st.markdown("## Step 3: Analyze Coverage")

    client = st.session_state.selected_client
    campaign = st.session_state.selected_campaign
    articles = st.session_state.articles

    st.markdown(f"**Client:** {client['name']} | **Campaign:** {campaign['name']}")
    st.markdown(f"**Articles to analyze:** {len(articles)}")

    # Check for API key
    api_key = st.secrets.get("anthropic_api_key", "")
    if not api_key:
        st.error("Anthropic API key not configured. Add it to .streamlit/secrets.toml")
        return

    col1, col2 = st.columns(2)
    with col1:
        if st.button("← Back"):
            st.session_state.step = 2
            st.rerun()
    with col2:
        if st.button("🚀 Start Analysis", type="primary"):
            run_analysis(api_key, client, campaign, articles)


def run_analysis(api_key: str, client: dict, campaign: dict, articles: list):
    """Run the full analysis pipeline."""
    results = []
    failed_articles = []

    progress_bar = st.progress(0)
    status_text = st.empty()

    anthropic_client = create_client(api_key)
    outlets_df = st.session_state.outlets_df

    total = len(articles)

    for i, article in enumerate(articles):
        progress = (i + 1) / total
        progress_bar.progress(progress)
        status_text.markdown(f"**Scraping:** {article.get('headline', article.get('url', ''))[:50]}...")

        # Step 1: Scrape content (or use override if provided from single article flow)
        if article.get("content_override"):
            # Single article flow - content already fetched/provided
            scrape_result = {
                "success": True,
                "content": article["content_override"],
                "title": article.get("headline", ""),
                "domain": article.get("domain", extract_domain(article.get("url", ""))),
                "word_count": len(article["content_override"].split()),
                "raw_html": article.get("raw_html"),
                "author": article.get("author"),
            }
        else:
            scrape_result = scrape_url(article["url"])

        if not scrape_result["success"]:
            failed_articles.append({
                "article": article,
                "error": scrape_result["error"],
            })
            continue

        # Step 2: Look up outlet (or use override if provided)
        outlet_info = lookup_outlet(scrape_result["domain"], outlets_df)
        if not outlet_info["found"]:
            outlet_info["name"] = article.get("publication", "Unknown")

        # Apply tier override if provided (from single article flow)
        if article.get("tier_override"):
            outlet_info["tier"] = article["tier_override"]

        status_text.markdown(f"**Scoring:** {article.get('headline', '')[:50]}...")

        # Step 3: Score with AI
        score_result = score_article(
            client=anthropic_client,
            article_content=scrape_result["content"],
            article_headline=article.get("headline", scrape_result.get("title", "")),
            outlet_name=outlet_info["name"],
            outlet_type=outlet_info.get("type", "Online"),
            outlet_tier=outlet_info.get("tier", 3),
            client_name=client["name"],
            industry=client.get("industry", ""),
            spokespeople=get_spokespeople(campaign),
            key_messages_high=campaign.get("key_messages_high", []),
            key_messages_medium=campaign.get("key_messages_medium", []),
            key_messages_low=campaign.get("key_messages_low", []),
            competitors=campaign.get("competitors", []),
            article_url=article.get("url", ""),
            raw_html=scrape_result.get("raw_html"),
            author_meta=scrape_result.get("author"),
        )

        results.append({
            "article": article,
            "scrape": scrape_result,
            "outlet": outlet_info,
            "scores": score_result,
        })

        time.sleep(0.3)

    progress_bar.progress(1.0)

    # Store results and go to step 4
    st.session_state.results = results
    st.session_state.failed_articles = failed_articles

    if failed_articles:
        status_text.markdown(f"**Analysis complete. {len(failed_articles)} articles could not be scraped - you can add them manually on the results page.**")
    else:
        status_text.markdown("**Analysis complete!**")

    # Always go to results page
    st.session_state.step = 4
    time.sleep(1)
    st.rerun()


def render_failed_articles_section(failed_articles: list, client: dict, campaign: dict):
    """Render section for manually adding failed articles with all required fields."""
    st.markdown("### ⚠️ Articles That Could Not Be Scraped")
    st.markdown(f"**{len(failed_articles)} articles** failed to scrape automatically. You can manually enter their details below to include them in scoring.")

    # Check for API key
    api_key = st.secrets.get("anthropic_api_key", "")
    if not api_key:
        st.error("Anthropic API key not configured.")
        return

    outlets_df = st.session_state.outlets_df

    for i, failed in enumerate(failed_articles):
        article = failed["article"]
        original_url = article.get("url", "")
        original_headline = article.get("headline", "")

        with st.expander(f"❌ {original_headline[:60] if original_headline else original_url[:60]}... - {failed['error']}", expanded=True):
            st.markdown(f"**Original URL:** {original_url}")
            st.caption(f"**Error:** {failed['error']}")

            st.markdown("---")
            st.markdown("**Enter article details manually:**")

            col1, col2 = st.columns(2)

            with col1:
                headline = st.text_input(
                    "Headline *",
                    value=original_headline,
                    key=f"manual_headline_{i}",
                    help="The article headline (used for brand-in-headline scoring)"
                )

                outlet_name = st.text_input(
                    "Outlet/Publication Name *",
                    value=article.get("publication", ""),
                    key=f"manual_outlet_{i}",
                    help="e.g., 'The Globe and Mail', 'TechCrunch' (used for tier lookup)"
                )

            with col2:
                byline = st.text_input(
                    "Author/Byline (optional)",
                    key=f"manual_byline_{i}",
                    help="The journalist's name, if known"
                )

                outlet_tier = st.selectbox(
                    "Outlet Tier (if known)",
                    options=["Auto-detect", "Tier 1", "Tier 2", "Tier 3"],
                    key=f"manual_tier_{i}",
                    help="Leave as auto-detect to look up from database, or manually set"
                )

            content = st.text_area(
                "Article Text *",
                key=f"manual_content_{i}",
                height=250,
                help="Paste the full article text here"
            )

            if st.button(f"📊 Analyze This Article", key=f"analyze_manual_{i}", type="primary"):
                if not headline:
                    st.error("Please enter a headline.")
                elif not outlet_name:
                    st.error("Please enter the outlet/publication name.")
                elif not content:
                    st.error("Please paste the article content.")
                else:
                    with st.spinner("Analyzing article..."):
                        # Look up outlet or use manual tier
                        outlet_info = lookup_outlet(extract_domain(original_url), outlets_df)

                        # If not found or user specified tier, use manual values
                        if not outlet_info["found"] or outlet_tier != "Auto-detect":
                            outlet_info["name"] = outlet_name
                            if outlet_tier != "Auto-detect":
                                outlet_info["tier"] = int(outlet_tier.split()[1])
                            else:
                                outlet_info["tier"] = 3  # Default
                            outlet_info["type"] = "Online"

                        # Create API client and score
                        anthropic_client = create_client(api_key)

                        score_result = score_article(
                            client=anthropic_client,
                            article_content=content,
                            article_headline=headline,
                            outlet_name=outlet_info["name"],
                            outlet_type=outlet_info.get("type", "Online"),
                            outlet_tier=outlet_info.get("tier", 3),
                            client_name=client["name"],
                            industry=client.get("industry", ""),
                            spokespeople=get_spokespeople(campaign),
                            key_messages_high=campaign.get("key_messages_high", []),
                            key_messages_medium=campaign.get("key_messages_medium", []),
                            key_messages_low=campaign.get("key_messages_low", []),
                            competitors=campaign.get("competitors", []),
                            article_url=original_url,
                            raw_html=None,  # Manual entry, no HTML available
                            author_meta=None,
                        )

                        # Add to results
                        new_result = {
                            "article": {
                                "url": original_url,
                                "headline": headline,
                                "publication": outlet_name,
                                "byline": byline,
                            },
                            "scrape": {"success": True, "content": content, "manual": True},
                            "outlet": outlet_info,
                            "scores": score_result,
                        }

                        st.session_state.results.append(new_result)

                        # Remove from failed list
                        st.session_state.failed_articles = [
                            f for f in st.session_state.failed_articles
                            if f["article"].get("url") != original_url
                        ]

                        if score_result.get("success"):
                            st.success(f"✅ Article analyzed! Score: **{score_result.get('total_score', 0)}** ({score_result.get('grade', 'N/A')})")
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error(f"Scoring failed: {score_result.get('error', 'Unknown error')}")

    st.markdown("---")


def step4_results():
    """Step 4: Display results dashboard."""
    st.markdown("## Results Dashboard")

    results = st.session_state.results
    failed_articles = st.session_state.failed_articles
    campaign = st.session_state.selected_campaign
    client = st.session_state.selected_client

    # Handle failed articles first - show at top if any exist
    if failed_articles:
        render_failed_articles_section(failed_articles, client, campaign)

    if not results and not failed_articles:
        st.warning("No results to display.")
        if st.button("← Start Over"):
            st.session_state.step = 1
            st.rerun()
        return

    if not results:
        st.info("No articles have been scored yet. Add articles manually above, or start a new analysis.")
        if st.button("← Start New Analysis"):
            st.session_state.step = 1
            st.rerun()
        return

    # Summary metrics
    successful = [r for r in results if r["scores"].get("success")]
    scores = [r["scores"]["total_score"] for r in successful]

    # Top row: summary stats
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        avg_score = sum(scores) / len(scores) if scores else 0
        avg_grade = get_score_grade(avg_score)
        st.markdown(f"<div class='big-score'>{avg_score:.0f}</div>", unsafe_allow_html=True)
        st.markdown(f"**Average Score** ({avg_grade})")

    with col2:
        st.metric("Articles Analyzed", len(successful))

    with col3:
        st.metric("Failed to Scrape", len(failed_articles))

    with col4:
        high_performers = len([s for s in scores if s >= 70])
        st.metric("High Performers (70+)", high_performers)

    st.markdown("---")

    # Tier breakdown summary
    st.markdown("### Average Score by Tier")
    render_tier_summary(successful)

    st.markdown("---")

    # Individual articles
    st.markdown("### Individual Articles")

    sort_by = st.selectbox(
        "Sort by:",
        options=["Score (High to Low)", "Score (Low to High)", "Outlet Tier"],
    )

    sorted_results = sort_results(successful, sort_by)

    for result in sorted_results:
        render_article_card(result)

    # Failed articles section
    if failed_articles:
        st.markdown("---")
        st.markdown("### ⚠️ Articles Not Analyzed")
        for failed in failed_articles:
            article = failed["article"]
            st.markdown(f"<div class='failed-article'>❌ **{article.get('headline', 'Unknown')}** - {failed['error']}</div>", unsafe_allow_html=True)

    st.markdown("---")

    # Export
    st.markdown("### Export Results")

    # PDF accent color (optional)
    with st.expander("PDF Options"):
        accent_color = st.color_picker(
            "Client Accent Color",
            value="#000000",
            help="Used sparingly in the PDF for borders and accents"
        )

    col1, col2, col3 = st.columns(3)

    with col1:
        excel_data = generate_excel_export(results, client, campaign)
        st.download_button(
            "📥 Download Excel",
            data=excel_data,
            file_name=f"coverage_scores_{client['name']}_{campaign['name']}_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with col2:
        # Generate PDF report
        logo_path = os.path.join(os.path.dirname(__file__), "assets", "CI-report-logo.png")
        if not os.path.exists(logo_path):
            logo_path = None

        pdf_data = generate_coverage_report(
            client_name=client["name"],
            campaign_name=campaign["name"],
            results=results,
            accent_color=accent_color,
            logo_path=logo_path,
        )
        st.download_button(
            "📄 Download PDF Report",
            data=pdf_data,
            file_name=f"coverage_report_{client['name']}_{campaign['name']}_{pd.Timestamp.now().strftime('%Y%m%d')}.pdf",
            mime="application/pdf",
        )

    with col3:
        if st.button("🔄 Start New Analysis"):
            st.session_state.step = 1
            st.session_state.articles = []
            st.session_state.results = []
            st.session_state.failed_articles = []
            st.rerun()


def render_tier_summary(results: list):
    """Render average scores per tier."""
    if not results:
        return

    tier_keys = [
        ("tier_1_foundational", "Foundational", 36),
        ("tier_2_messaging", "Messaging", 30),
        ("tier_3_prominence", "Prominence", 10),
        ("tier_4_credibility", "Credibility", 10),
        ("tier_5_competitive", "Competitive", 9),
        ("tier_6_audience_fit", "Audience Fit", 6.5),
        ("tier_7_supporting", "Supporting", 2.5),
        ("bonus", "Bonus", 16),
    ]

    cols = st.columns(4)

    for i, (tier_key, tier_name, max_pts) in enumerate(tier_keys):
        col = cols[i % 4]

        # Calculate average for this tier
        tier_scores = []
        for r in results:
            if r["scores"].get("tier_scores") and tier_key in r["scores"]["tier_scores"]:
                tier_scores.append(r["scores"]["tier_scores"][tier_key]["score"])

        avg = sum(tier_scores) / len(tier_scores) if tier_scores else 0
        pct = (avg / max_pts * 100) if max_pts > 0 else 0

        with col:
            color = get_score_color(pct)
            st.markdown(f"**{tier_name}**")
            st.markdown(f"<span style='font-size:20px; color:{color};'>{avg:.1f}</span> / {max_pts}", unsafe_allow_html=True)
            st.progress(min(1.0, pct / 100))


def sort_results(results: list, sort_by: str) -> list:
    """Sort results based on user selection."""
    if sort_by == "Score (High to Low)":
        return sorted(results, key=lambda x: x["scores"].get("total_score", 0), reverse=True)
    elif sort_by == "Score (Low to High)":
        return sorted(results, key=lambda x: x["scores"].get("total_score", 0))
    elif sort_by == "Outlet Tier":
        return sorted(results, key=lambda x: x["outlet"].get("tier", 3))
    return results


def render_article_card(result: dict):
    """Render a single article result card with tier breakdown."""
    article = result["article"]
    scores = result["scores"]
    outlet = result["outlet"]

    if not scores.get("success"):
        with st.expander(f"❌ {article.get('headline', 'Unknown')[:60]} - Scoring failed"):
            st.error(scores.get("error", "Unknown error"))
        return

    total = scores.get("total_score", 0)
    grade = scores.get("grade", "N/A")
    color = get_score_color(total)

    # Build expander title with op-ed indicator
    headline = article.get('headline', 'Unknown')[:60]
    op_ed_badge = "✍️ " if scores.get("is_client_op_ed") else ""

    # Header with score
    with st.expander(f"**{grade}** ({total:.0f}) - {op_ed_badge}{headline}"):
        col1, col2 = st.columns([1, 3])

        with col1:
            st.markdown(f"<div style='font-size:48px; font-weight:bold; color:{color};'>{total:.0f}</div>", unsafe_allow_html=True)
            st.markdown(f"**Grade: {grade}**")
            st.markdown(f"**Outlet:** {outlet.get('name', 'Unknown')}")
            st.markdown(f"**Tier:** {outlet.get('tier', 'N/A')} | **Type:** {outlet.get('type', 'N/A')}")

            # Show article link
            article_url = article.get('url', '')
            if article_url:
                st.markdown(f"[📄 Read Article]({article_url})")

            # Show op-ed indicator if detected
            if scores.get("is_client_op_ed"):
                st.markdown(f"<span style='background:#10B981; color:white; padding:2px 8px; border-radius:4px; font-size:12px;'>✍️ CLIENT OP-ED</span>", unsafe_allow_html=True)
                if scores.get("matched_spokesperson"):
                    st.markdown(f"*By {scores.get('matched_spokesperson')}*")

        with col2:
            # Tier breakdown
            st.markdown("**Tier Breakdown:**")

            tier_scores = scores.get("tier_scores", {})

            for tier_key, tier_data in tier_scores.items():
                tier_name = tier_data.get("name", tier_key)
                tier_score = tier_data.get("score", 0)
                tier_max = tier_data.get("max", 0)
                pct = (tier_score / tier_max * 100) if tier_max > 0 else 0

                # Show penalties in red
                if tier_score < 0:
                    st.markdown(f"<span class='penalty-tag'>{tier_name}: {tier_score:.1f}/{tier_max}</span>", unsafe_allow_html=True)
                else:
                    st.markdown(f"{tier_name}: **{tier_score:.1f}**/{tier_max}")

        # Summary
        st.markdown("---")
        st.markdown(f"**Summary:** {scores.get('summary', 'No summary available.')}")

        # Key messages found
        messages_found = scores.get("key_messages_found", [])
        if messages_found:
            st.markdown(f"**Key Messages Found:** {', '.join(messages_found[:3])}{'...' if len(messages_found) > 3 else ''}")

        # Competitors found
        competitors_found = scores.get("competitors_found", [])
        if competitors_found:
            st.markdown(f"<span class='penalty-tag'>**Competitors Mentioned:** {', '.join(competitors_found)}</span>", unsafe_allow_html=True)

        # Detailed scores expander
        with st.expander("View all factor scores"):
            raw_scores = scores.get("scores", {})
            for tier_key, tier_factors in raw_scores.items():
                if isinstance(tier_factors, dict) and tier_key not in ["summary", "key_messages_found", "competitors_found"]:
                    # Get max points for this tier from SCORING_MODEL
                    tier_info = SCORING_MODEL.get(tier_key, {})
                    tier_max = tier_info.get("max_points", 0)
                    factor_maxes = {k: v.get("max", 0) for k, v in tier_info.get("factors", {}).items()}

                    st.markdown(f"**{tier_info.get('name', tier_key.replace('_', ' ').title())}:**")
                    for factor, factor_data in tier_factors.items():
                        # Handle new format with explanations
                        if isinstance(factor_data, dict) and "score" in factor_data:
                            score = factor_data["score"]
                            explanation = factor_data.get("explanation", "")
                            factor_max = factor_maxes.get(factor, "?")

                            if score < 0:
                                st.markdown(f"  - **{factor.replace('_', ' ').title()}**: <span style='color:#EF4444;font-weight:bold;'>{score}</span> — _{explanation}_", unsafe_allow_html=True)
                            else:
                                st.markdown(f"  - **{factor.replace('_', ' ').title()}**: {score}/{factor_max} — _{explanation}_")
                        # Handle old format (just numbers)
                        elif isinstance(factor_data, (int, float)):
                            factor_max = factor_maxes.get(factor, "?")
                            if factor_data < 0:
                                st.markdown(f"  - **{factor.replace('_', ' ').title()}**: <span style='color:#EF4444;font-weight:bold;'>{factor_data}</span>", unsafe_allow_html=True)
                            else:
                                st.markdown(f"  - **{factor.replace('_', ' ').title()}**: {factor_data}/{factor_max}")


def generate_excel_export(results: list, client: dict, campaign: dict) -> bytes:
    """Generate Excel export of results."""
    rows = []

    for result in results:
        article = result["article"]
        scores = result["scores"]
        outlet = result["outlet"]

        row = {
            "Headline": article.get("headline", ""),
            "URL": article.get("url", ""),
            "Publication": article.get("publication", outlet.get("name", "")),
            "Date": article.get("date", ""),
            "Outlet Tier": outlet.get("tier", ""),
            "Outlet Type": outlet.get("type", ""),
            "Total Score": scores.get("total_score", "") if scores.get("success") else "Error",
            "Grade": scores.get("grade", ""),
        }

        # Add tier scores
        if scores.get("success") and scores.get("tier_scores"):
            for tier_key, tier_data in scores["tier_scores"].items():
                row[tier_data["name"]] = tier_data["score"]

        row["Summary"] = scores.get("summary", "")
        row["Key Messages Found"] = "; ".join(scores.get("key_messages_found", []))
        row["Competitors Found"] = "; ".join(scores.get("competitors_found", []))

        rows.append(row)

    df = pd.DataFrame(rows)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Coverage Scores")

        # Add campaign info sheet
        campaign_info = pd.DataFrame([{
            "Client": client["name"],
            "Campaign": campaign["name"],
            "Industry": client.get("industry", ""),
            "Spokespeople": ", ".join(get_spokespeople(campaign)),
            "High Priority Messages": "\n".join(campaign.get("key_messages_high", [])),
            "Medium Priority Messages": "\n".join(campaign.get("key_messages_medium", [])),
            "Low Priority Messages": "\n".join(campaign.get("key_messages_low", [])),
            "Competitors": ", ".join(campaign.get("competitors", [])),
        }])
        campaign_info.to_excel(writer, index=False, sheet_name="Campaign Info")

    return output.getvalue()


def main():
    """Main application entry point."""
    init_session_state()

    # Authentication
    if not check_auth():
        return

    # Load data
    load_data()

    # Main navigation using tabs
    tab1, tab2, tab3 = st.tabs(["📊 Analyze Coverage", "👥 Client Profiles", "🔧 Tools"])

    with tab1:
        render_analyze_tab()

    with tab2:
        render_client_profiles()

    with tab3:
        render_tools()


if __name__ == "__main__":
    main()
